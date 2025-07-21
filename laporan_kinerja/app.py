import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os
from datetime import datetime

FILE_NAME = "laporan_kinerja.xlsx"

def buat_file_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Februari 2025"
    header = ["No", "Hari/Tanggal", "DL/Dinas", "Rincian Aktivitas/Tugas", "Kuantitas", "Output", "Waktu (Menit)"]
    ws.append(header)
    wb.save(FILE_NAME)

def simpan_data():
    tanggal = ent_tanggal.get()
    jenis = ent_jenis.get()
    aktivitas = ent_aktivitas.get("1.0", "end").strip()
    kuantitas = int(ent_kuantitas.get())
    output = ent_output.get()
    waktu = int(ent_waktu.get())

    if not os.path.exists(FILE_NAME):
        buat_file_excel()

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    no = ws.max_row  # Auto-numbering
    hari = datetime.strptime(tanggal, "%Y-%m-%d").strftime("%A, %d %B %Y")
    ws.append([no, hari, jenis, aktivitas, kuantitas, output, waktu])
    wb.save(FILE_NAME)

    messagebox.showinfo("Sukses", "Data berhasil disimpan.")
    ent_aktivitas.delete("1.0", "end")

def hitung_total():
    if not os.path.exists(FILE_NAME):
        messagebox.showwarning("Gagal", "Belum ada data.")
        return

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    total_menit = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and isinstance(row[6], int):
            total_menit += row[6]

    npk = total_menit / 45.6
    ck = npk / 2 * 100

    hasil = f"""
Jumlah Waktu Penyelesaian (JWP): {total_menit} menit
Nilai Produktivitas Kerja (NPK): {npk:.2f}
Capaian Kinerja (CK): {ck:.2f}
    """
    messagebox.showinfo("Hasil Perhitungan", hasil)

# GUI
root = tk.Tk()
root.title("Laporan Kinerja Bulanan")

tk.Label(root, text="Tanggal (YYYY-MM-DD)").grid(row=0, column=0)
ent_tanggal = tk.Entry(root)
ent_tanggal.grid(row=0, column=1)

tk.Label(root, text="DL/Dinas").grid(row=1, column=0)
ent_jenis = tk.Entry(root)
ent_jenis.insert(0, "DL")
ent_jenis.grid(row=1, column=1)

tk.Label(root, text="Rincian Aktivitas").grid(row=2, column=0)
ent_aktivitas = tk.Text(root, height=4, width=40)
ent_aktivitas.grid(row=2, column=1)

tk.Label(root, text="Kuantitas").grid(row=3, column=0)
ent_kuantitas = tk.Entry(root)
ent_kuantitas.insert(0, "1")
ent_kuantitas.grid(row=3, column=1)

tk.Label(root, text="Output").grid(row=4, column=0)
ent_output = tk.Entry(root)
ent_output.insert(0, "kegiatan")
ent_output.grid(row=4, column=1)

tk.Label(root, text="Waktu (menit)").grid(row=5, column=0)
ent_waktu = tk.Entry(root)
ent_waktu.insert(0, "480")
ent_waktu.grid(row=5, column=1)

tk.Button(root, text="Simpan", command=simpan_data).grid(row=6, column=0, pady=10)
tk.Button(root, text="Hitung NPK & CK", command=hitung_total).grid(row=6, column=1, pady=10)

root.mainloop()
